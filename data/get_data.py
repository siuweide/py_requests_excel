import requests

from util.operation_excel import OperationExcel
from util.operation_json import OperationJson
from openpyxl import load_workbook
from data import dataconfig
import json

class GetData():

    def __init__(self):
        self.opera_excel = OperationExcel()
        self.opera_json = OperationJson()
        self.col = dataconfig

    def get_id(self, row):
        """ 获取id数据 """
        col = self.col.global_id()
        id = self.opera_excel.get_cell_value(row, col)
        return id

    def get_url(self, row):
        """ 获取url数据 """
        col = self.col.global_url()
        url = self.opera_excel.get_cell_value(row, col)
        return url

    def get_is_run(self, row):
        """ 获取是否运行用例 """
        res = None
        col = self.col.global_is_run()
        is_run = self.opera_excel.get_cell_value(row, col)
        if is_run == 'yes':
            res = True
        elif is_run == 'no':
            res = False
        return res

    def get_method(self, row):
        """ 获取请求方法 """
        col = self.col.global_method()
        method = self.opera_excel.get_cell_value(row, col)
        return method

    def is_cookie(self, row):
        """ 判断是否获取cookie """
        col = self.col.global_headers()
        cookie = self.opera_excel.get_cell_value(row, col)
        return cookie

    def get_dependent_id(self, row):
        """ 获取依赖ID """
        col = self.col.global_dependent_id()
        dependent_id = self.opera_excel.get_cell_value(row, col)
        return dependent_id

    def get_dependent_data(self, row):
        """ 获取依赖数据 """
        col = self.col.global_dependent_data()
        dependent_data = self.opera_excel.get_cell_value(row, col)
        return dependent_data

    def get_dependent_field(self, row):
        """ 获取依赖字段 """
        col = self.col.global_dependent_field()
        dependent_field = self.opera_excel.get_cell_value(row, col)
        return dependent_field

    def get_headers_data(self, row):
        """ 获取请求数据 """
        res = ''
        col = self.col.global_data()
        data = self.opera_excel.get_cell_value(row, col)
        if data != '':
            res = self.opera_json.get_data(data)
        return res

    def get_expected_result(self, row):
        """ 获取预期结果 """
        col = self.col.global_expected_result()
        data = self.opera_excel.get_cell_value(row,col)
        return data

    def write_actual_result(self, row, message, file_name='../dataconfig/test1.xlsx'):
        """ 写入实际结果 """
        col = self.col.global_actual_result()
        # 生成一个已存在的对象
        wb = load_workbook(file_name)
        # 激活shell
        active_sheet = wb.active
        # 往shell中写入数据
        active_sheet.cell(row, col+1, message)
        # 保存结果
        wb.save(file_name)

    def write_json_cookie(self, case, file_name='../dataconfig/user_json_cookie.json'):
        # 将cookie写入到文件当中
        cookies = requests.utils.dict_from_cookiejar(case.cookies)
        cookie = "{\"Cookie\":\"PHPSESSID=%s\"}" % cookies['PHPSESSID']
        with open(file_name, 'w') as f:
            f.write(cookie)

    def read_cookie(self, file_name='../dataconfig/user_json_cookie.json'):
        # 读取cookie文件
        with open(file_name) as f:
            cookie = f.read()
            return json.loads(cookie)

if __name__ == '__main__':
    test = GetData()
    print(test.read_cookie())